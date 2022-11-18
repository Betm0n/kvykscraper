import openpyxl
from scraper import *



def filemod(filenames):
    driver = initiatedriver()
    for i,filename in enumerate(filenames):
        path = "C:/Users/kvyk/Downloads/" + filename + f"{i}.xlsx"

        wb = openpyxl.load_workbook(path)
        sheet = wb.active 

        for i in range(1, sheet.max_row):
            cell = sheet.cell(row = i+1, column = 1)
            amazonprice = pricecheckamazon(driver, cell.value)
            flipkartprice = pricecheckflipkart(driver, cell.value)
            snapdealprice = pricechecksnapdeal(driver, cell.value)
            cell = sheet.cell(row = i+1, column = 2)
            cell.value = amazonprice
            cell = sheet.cell(row = i+1, column = 3)
            cell.value = flipkartprice
            cell = sheet.cell(row = i+1, column = 4)
            cell.value = snapdealprice

            wb.save(path)
    driverquit(driver)
        
    """
    assumes excel file format as:
    
        ProductName    Amazon   Flipkart    SnapDeal
            abc          -         -           -
            xyz          -         -           -
    
    """